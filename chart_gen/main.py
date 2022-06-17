absolute_path = 'C:/Users/Joon/PycharmProjects/chart_gen/'
save_path = 'C:/Users/Joon/PycharmProjects/chart_gen/_save/'

from openpyxl import load_workbook
from openpyxl.styles import Font
#불러오기
wb = load_workbook(absolute_path+'form.xlsx')
ws = wb['Sheet1']

#값 채우기
pat_num=int(input('환자 번호 입력'))
pat_name=(input('환자 이름 입력'))
# ans_list = ['ans1', 'ans2', 'ans3', '아니오', '아니오', '예', '아니오', 'ans8', 'ans9', 'ans10', '당뇨병', '아니오', '예']
ans_list = ['팔꿈치', '7', '100도 이상', '예', '아니오', '예', '아니오', '일주일에 4회 이상', '일주일에 2~3회', '월 1회', '당뇨병', '아니오', '예']
ans_form = ['B6', 'B7', 'B8', 'B9', 'B10', 'B11', 'B12', 'E6', 'E7', 'E8', 'E9', 'E10', 'E11']
ws['D3'] = pat_num
ws['E3'] = pat_name
for i in range(len(ans_list)):
    ws[ans_form[i]] = ans_list[i]

#질문 4~7 빨간줄, 사회력 5, 6 빨간줄, 사회력 4 답변 빨간줄
#답변 강조
hi_list1 = [3, 4, 5, 6]
hi_list1_ = ['A9', 'A10', 'A11', 'A12']
hi_list2 = [11, 12]
hi_list2_ = ['D10', 'D11']
for i in range(len(hi_list1)):
    if ans_list[hi_list1[i]] == '아니오':
        ws[hi_list1_[i]].font = Font(color='00FF0000')

for i in range(len(hi_list2)):
    if ans_list[hi_list2[i]] == '예':
        ws[hi_list2_[i]].font = Font(color='00FF0000')

if ans_list[10] != '아니오':
    ws['D9'].font = Font(color='00FF0000')

wb.save(save_path + 'report{}{}.xlsx'.format(pat_num, pat_name))